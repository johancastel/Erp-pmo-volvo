# src/database/repositories/novedades_repository.py
import datetime
import pandas as pd
from src.config.settings import PATIOS_CONFIG
from src.database.connection import get_connection_manager
from src.database.queries.sql_templates import (
    UNIFIED_COLS,
    CREATE_VEHICULOS_SQLITE, CREATE_VEHICULOS_MYSQL,
    CREATE_INSPECCIONES_SQLITE, CREATE_INSPECCIONES_MYSQL,
    CREATE_NIVELES_SQLITE, CREATE_NIVELES_MYSQL,
    CREATE_MECANICA_SQLITE, CREATE_MECANICA_MYSQL,
    CREATE_ELECTRICA_SQLITE, CREATE_ELECTRICA_MYSQL,
    CREATE_NOVEDADES_VOLVO_SQLITE, CREATE_NOVEDADES_VOLVO_MYSQL
)

class NovedadesRepository:
    def __init__(self, connection_manager=None):
        self.connection_manager = connection_manager or get_connection_manager()

    def init_db(self, host, user, password, db_name, port):
        """
        Initializes the normalized database structure and performs automatic migration
        from the legacy 'novedades_moviles' table if it exists.
        """
        conn, db_type = self.connection_manager.get_connection(host, user, password, db_name, port)
        cursor = conn.cursor()

        # 1. Create new normalized tables
        if db_type == "MySQL":
            cursor.execute(CREATE_VEHICULOS_MYSQL)
            cursor.execute(CREATE_INSPECCIONES_MYSQL)
            cursor.execute(CREATE_NIVELES_MYSQL)
            cursor.execute(CREATE_MECANICA_MYSQL)
            cursor.execute(CREATE_ELECTRICA_MYSQL)
            cursor.execute(CREATE_NOVEDADES_VOLVO_MYSQL)
        else:
            cursor.execute(CREATE_VEHICULOS_SQLITE)
            cursor.execute(CREATE_INSPECCIONES_SQLITE)
            cursor.execute(CREATE_NIVELES_SQLITE)
            cursor.execute(CREATE_MECANICA_SQLITE)
            cursor.execute(CREATE_ELECTRICA_SQLITE)
            cursor.execute(CREATE_NOVEDADES_VOLVO_SQLITE)
        conn.commit()

        # Ensure new columns exist in novedades_volvo
        try:
            if db_type == "MySQL":
                # Check for tecnico_correccion
                cursor.execute("""
                    SELECT COLUMN_NAME 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_NAME = 'novedades_volvo' 
                      AND TABLE_SCHEMA = %s 
                      AND COLUMN_NAME = 'tecnico_correccion'
                """, (db_name,))
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE novedades_volvo ADD COLUMN tecnico_correccion VARCHAR(100) DEFAULT ''")
                
                # Check for insumos_usados
                cursor.execute("""
                    SELECT COLUMN_NAME 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_NAME = 'novedades_volvo' 
                      AND TABLE_SCHEMA = %s 
                      AND COLUMN_NAME = 'insumos_usados'
                """, (db_name,))
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE novedades_volvo ADD COLUMN insumos_usados TEXT")
                    
                # Check for cantidad
                cursor.execute("""
                    SELECT COLUMN_NAME 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_NAME = 'novedades_volvo' 
                      AND TABLE_SCHEMA = %s 
                      AND COLUMN_NAME = 'cantidad'
                """, (db_name,))
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE novedades_volvo ADD COLUMN cantidad VARCHAR(50) DEFAULT ''")
            else:
                try:
                    cursor.execute("ALTER TABLE novedades_volvo ADD COLUMN tecnico_correccion TEXT DEFAULT ''")
                except Exception:
                    pass
                try:
                    cursor.execute("ALTER TABLE novedades_volvo ADD COLUMN insumos_usados TEXT DEFAULT ''")
                except Exception:
                    pass
                try:
                    cursor.execute("ALTER TABLE novedades_volvo ADD COLUMN cantidad TEXT DEFAULT ''")
                except Exception:
                    pass
            conn.commit()
        except Exception:
            pass

        # 2. Check for legacy table 'novedades_moviles' to migrate data
        legacy_table_exists = False
        try:
            cursor.execute("SELECT 1 FROM novedades_moviles LIMIT 1")
            cursor.fetchall()
            legacy_table_exists = True
        except Exception:
            legacy_table_exists = False

        if legacy_table_exists:
            print("Legacy 'novedades_moviles' table detected. Starting automatic data migration...")
            try:
                # Read all legacy data
                cursor.execute("""
                    SELECT 
                        patio, movil,
                        nivel_aceite_motor_si_no, nivel_aceite_motor_cant,
                        nivel_refrigerante_si_no, nivel_refrigerante_cant,
                        nivel_aceite_hidraulico_si_no, nivel_aceite_hidraulico_cant,
                        nivel_limpiabrisas_si_no, nivel_limpiabrisas_cant,
                        inspeccion_ducto_admision, drenar_tanques, observacion_niveles,
                        revision_fugas, separador_humedad, inspeccion_embrague,
                        inspeccion_palanca, inspeccion_ventilador, inspeccion_suspension,
                        inspeccion_frenos, inspeccion_direccion, elec_luces,
                        elec_tablero, elec_rutero, elec_arranque, elec_puertas,
                        elec_baterias, detalles, observaciones, detalles_elec,
                        observaciones_elec, insumos, tecnico, fecha_registro
                    FROM novedades_moviles
                """)
                legacy_rows = cursor.fetchall()
                
                # Migrate records one by one
                for r in legacy_rows:
                    # Convert row to dictionary/tuple index independent format
                    if isinstance(r, dict):
                        row_dict = r
                    else:
                        row_dict = {
                            'patio': r[0], 'movil': r[1],
                            'nivel_aceite_motor_si_no': r[2], 'nivel_aceite_motor_cant': r[3],
                            'nivel_refrigerante_si_no': r[4], 'nivel_refrigerante_cant': r[5],
                            'nivel_aceite_hidraulico_si_no': r[6], 'nivel_aceite_hidraulico_cant': r[7],
                            'nivel_limpiabrisas_si_no': r[8], 'nivel_limpiabrisas_cant': r[9],
                            'inspeccion_ducto_admision': r[10], 'drenar_tanques': r[11], 'observacion_niveles': r[12],
                            'revision_fugas': r[13], 'separador_humedad': r[14], 'inspeccion_embrague': r[15],
                            'inspeccion_palanca': r[16], 'inspeccion_ventilador': r[17], 'inspeccion_suspension': r[18],
                            'inspeccion_frenos': r[19], 'inspeccion_direccion': r[20], 'elec_luces': r[21],
                            'elec_tablero': r[22], 'elec_rutero': r[23], 'elec_arranque': r[24], 'elec_puertas': r[25],
                            'elec_baterias': r[26], 'detalles': r[27], 'observaciones': r[28], 'detalles_elec': r[29],
                            'observaciones_elec': r[30], 'insumos': r[31], 'tecnico': r[32], 'fecha_registro': r[33]
                        }
                    
                    # 1. Insert vehicle into vehiculos table
                    if db_type == "MySQL":
                        cursor.execute("""
                            INSERT IGNORE INTO vehiculos (movil, patio_predeterminado, estado)
                            VALUES (%s, %s, 'ACTIVO')
                        """, (row_dict['movil'], row_dict['patio']))
                    else:
                        cursor.execute("""
                            INSERT OR IGNORE INTO vehiculos (movil, patio_predeterminado, estado)
                            VALUES (?, ?, 'ACTIVO')
                        """, (row_dict['movil'], row_dict['patio']))

                    # 2. Insert main inspection record
                    tecnico_val = row_dict['tecnico'] if row_dict['tecnico'] else ""
                    fecha_val = row_dict['fecha_registro'] if row_dict['fecha_registro'] else ""
                    
                    if db_type == "MySQL":
                        cursor.execute("""
                            INSERT INTO inspecciones (movil, patio, tecnico, fecha_registro, tipo_inspeccion)
                            VALUES (%s, %s, %s, %s, 'COMBINADA')
                        """, (row_dict['movil'], row_dict['patio'], tecnico_val, fecha_val))
                        inspeccion_id = cursor.lastrowid
                    else:
                        cursor.execute("""
                            INSERT INTO inspecciones (movil, patio, tecnico, fecha_registro, tipo_inspeccion)
                            VALUES (?, ?, ?, ?, 'COMBINADA')
                        """, (row_dict['movil'], row_dict['patio'], tecnico_val, fecha_val))
                        inspeccion_id = cursor.lastrowid

                    # 3. Insert child inspeccion_niveles record
                    if db_type == "MySQL":
                        cursor.execute("""
                            INSERT INTO inspeccion_niveles (
                                inspeccion_id, nivel_aceite_motor_si_no, nivel_aceite_motor_cant,
                                nivel_refrigerante_si_no, nivel_refrigerante_cant,
                                nivel_aceite_hidraulico_si_no, nivel_aceite_hidraulico_cant,
                                nivel_limpiabrisas_si_no, nivel_limpiabrisas_cant,
                                inspeccion_ducto_admision, drenar_tanques, observacion_niveles
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            inspeccion_id, row_dict['nivel_aceite_motor_si_no'], row_dict['nivel_aceite_motor_cant'],
                            row_dict['nivel_refrigerante_si_no'], row_dict['nivel_refrigerante_cant'],
                            row_dict['nivel_aceite_hidraulico_si_no'], row_dict['nivel_aceite_hidraulico_cant'],
                            row_dict['nivel_limpiabrisas_si_no'], row_dict['nivel_limpiabrisas_cant'],
                            row_dict['inspeccion_ducto_admision'], row_dict['drenar_tanques'], row_dict['observacion_niveles']
                        ))
                    else:
                        cursor.execute("""
                            INSERT INTO inspeccion_niveles (
                                inspeccion_id, nivel_aceite_motor_si_no, nivel_aceite_motor_cant,
                                nivel_refrigerante_si_no, nivel_refrigerante_cant,
                                nivel_aceite_hidraulico_si_no, nivel_aceite_hidraulico_cant,
                                nivel_limpiabrisas_si_no, nivel_limpiabrisas_cant,
                                inspeccion_ducto_admision, drenar_tanques, observacion_niveles
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            inspeccion_id, row_dict['nivel_aceite_motor_si_no'], row_dict['nivel_aceite_motor_cant'],
                            row_dict['nivel_refrigerante_si_no'], row_dict['nivel_refrigerante_cant'],
                            row_dict['nivel_aceite_hidraulico_si_no'], row_dict['nivel_aceite_hidraulico_cant'],
                            row_dict['nivel_limpiabrisas_si_no'], row_dict['nivel_limpiabrisas_cant'],
                            row_dict['inspeccion_ducto_admision'], row_dict['drenar_tanques'], row_dict['observacion_niveles']
                        ))

                    # 4. Insert child inspeccion_mecanica record
                    if db_type == "MySQL":
                        cursor.execute("""
                            INSERT INTO inspeccion_mecanica (
                                inspeccion_id, revision_fugas, separador_humedad, inspeccion_embrague,
                                inspeccion_palanca, inspeccion_ventilador, inspeccion_suspension,
                                inspeccion_frenos, inspeccion_direccion, detalles, observaciones, insumos
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            inspeccion_id, row_dict['revision_fugas'], row_dict['separador_humedad'], row_dict['inspeccion_embrague'],
                            row_dict['inspeccion_palanca'], row_dict['inspeccion_ventilador'], row_dict['inspeccion_suspension'],
                            row_dict['inspeccion_frenos'], row_dict['inspeccion_direccion'], row_dict['detalles'], row_dict['observaciones'], row_dict['insumos']
                        ))
                    else:
                        cursor.execute("""
                            INSERT INTO inspeccion_mecanica (
                                inspeccion_id, revision_fugas, separador_humedad, inspeccion_embrague,
                                inspeccion_palanca, inspeccion_ventilador, inspeccion_suspension,
                                inspeccion_frenos, inspeccion_direccion, detalles, observaciones, insumos
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            inspeccion_id, row_dict['revision_fugas'], row_dict['separador_humedad'], row_dict['inspeccion_embrague'],
                            row_dict['inspeccion_palanca'], row_dict['inspeccion_ventilador'], row_dict['inspeccion_suspension'],
                            row_dict['inspeccion_frenos'], row_dict['inspeccion_direccion'], row_dict['detalles'], row_dict['observaciones'], row_dict['insumos']
                        ))

                    # 5. Insert child inspeccion_electrica record
                    if db_type == "MySQL":
                        cursor.execute("""
                            INSERT INTO inspeccion_electrica (
                                inspeccion_id, elec_luces, elec_tablero, elec_rutero,
                                elec_arranque, elec_puertas, elec_baterias, detalles_elec, observaciones_elec
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            inspeccion_id, row_dict['elec_luces'], row_dict['elec_tablero'], row_dict['elec_rutero'],
                            row_dict['elec_arranque'], row_dict['elec_puertas'], row_dict['elec_baterias'], row_dict['detalles_elec'], row_dict['observaciones_elec']
                        ))
                    else:
                        cursor.execute("""
                            INSERT INTO inspeccion_electrica (
                                inspeccion_id, elec_luces, elec_tablero, elec_rutero,
                                elec_arranque, elec_puertas, elec_baterias, detalles_elec, observaciones_elec
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            inspeccion_id, row_dict['elec_luces'], row_dict['elec_tablero'], row_dict['elec_rutero'],
                            row_dict['elec_arranque'], row_dict['elec_puertas'], row_dict['elec_baterias'], row_dict['detalles_elec'], row_dict['observaciones_elec']
                        ))

                conn.commit()

                # Rename the legacy table to backup
                now_sfx = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                if db_type == "MySQL":
                    cursor.execute(f"RENAME TABLE novedades_moviles TO novedades_moviles_old_backup_{now_sfx}")
                else:
                    cursor.execute(f"ALTER TABLE novedades_moviles RENAME TO novedades_moviles_old_backup_{now_sfx}")
                conn.commit()
                print("Automatic data migration finished successfully.")

            except Exception as e:
                conn.rollback()
                print(f"Failed to migrate legacy data: {str(e)}")

        # 3. Synchronize configured fleet vehicles in the 'vehiculos' table
        for patio_name, moviles_list in PATIOS_CONFIG.items():
            for movil_id in moviles_list:
                if db_type == "MySQL":
                    cursor.execute("""
                        INSERT IGNORE INTO vehiculos (movil, patio_predeterminado, estado)
                        VALUES (%s, %s, 'ACTIVO')
                    """, (movil_id, patio_name))
                else:
                    cursor.execute("""
                        INSERT OR IGNORE INTO vehiculos (movil, patio_predeterminado, estado)
                        VALUES (?, ?, 'ACTIVO')
                    """, (movil_id, patio_name))
        conn.commit()

        # 4. Migrate legacy/historical inspection mechanical & electrical observations to novedades_volvo
        try:
            # Check if table already has records to prevent slow re-migration
            cursor.execute("SELECT COUNT(*) AS total FROM novedades_volvo")
            count_res = cursor.fetchone()
            total_novedades = 0
            if count_res:
                if isinstance(count_res, dict):
                    total_novedades = count_res.get("total", 0)
                else:
                    total_novedades = count_res[0]
            
            if total_novedades > 0:
                print(f"Table 'novedades_volvo' already has {total_novedades} records. Skipping historical migration.")
                cursor.close()
                conn.close()
                return

            from src.services.export_service import ExportService
            fleet_info = ExportService.load_fleet_master()
            novedades_master = ExportService.load_novedades_master()
            
            cursor.execute("""
                SELECT 
                    i.movil, i.tecnico, i.fecha_registro, i.tipo_inspeccion,
                    m.detalles, m.observaciones,
                    e.detalles_elec, e.observaciones_elec
                FROM inspecciones i
                LEFT JOIN inspeccion_mecanica m ON i.id = m.inspeccion_id
                LEFT JOIN inspeccion_electrica e ON i.id = e.inspeccion_id
                WHERE i.tecnico != '' AND i.tecnico IS NOT NULL
            """)
            completed_rows = cursor.fetchall()
            
            for r in completed_rows:
                if isinstance(r, dict):
                    r_dict = {
                        'movil': r.get('movil', ''),
                        'tecnico': r.get('tecnico', ''),
                        'fecha_parsed': r.get('fecha_registro', ''),
                        'tipo_inspeccion': r.get('tipo_inspeccion', 'ALISTAMIENTO'),
                        'detalles_mec': r.get('detalles', ''),
                        'observaciones_mec': r.get('observaciones', ''),
                        'detalles_elec': r.get('detalles_elec', ''),
                        'observaciones_elec': r.get('observaciones_elec', '')
                    }
                else:
                    r_dict = {
                        'movil': r[0], 'tecnico': r[1], 'fecha_parsed': r[2], 'tipo_inspeccion': r[3],
                        'detalles_mec': r[4], 'observaciones_mec': r[5],
                        'detalles_elec': r[6], 'observaciones_elec': r[7]
                    }
                    
                movil_id = r_dict['movil']
                movil_clean = str(movil_id).strip()
                fleet_row = fleet_info.get(movil_clean, {})
                chasis_val = fleet_row.get("chasis", "")
                linea_val = fleet_row.get("linea", "")
                contrato_val = fleet_row.get("contrato", "")
                
                fecha_reg = r_dict['fecha_parsed'] if r_dict['fecha_parsed'] else ""
                tipo_ins = r_dict['tipo_inspeccion'] if r_dict['tipo_inspeccion'] else "ALISTAMIENTO"
                fecha_nov = fecha_reg[:10] if len(fecha_reg) > 10 else fecha_reg
                
                # Map source/fuente based on inspection type
                if tipo_ins in ["ALISTAMIENTO", "COMBINADA"]:
                    fuente = "ALISTAMIENTO"
                elif tipo_ins == "CORRECTIVO":
                    fuente = "MANTENIMIENTO CORRECTIVO"
                elif tipo_ins == "PREVENTIVO":
                    fuente = "MANTENIMIENTO PREVENTIVO"
                else:
                    fuente = f"MANTENIMIENTO {tipo_ins}" if tipo_ins else "ALISTAMIENTO"
                
                # 4.1 Mechanical Observation
                obs_mec = str(r_dict['observaciones_mec']).strip() if r_dict['observaciones_mec'] else ""
                det_mec = str(r_dict['detalles_mec']).strip() if r_dict['detalles_mec'] else ""
                if obs_mec and obs_mec.upper() not in ["N/A", "NONE", ""]:
                    if db_type == "MySQL":
                        cursor.execute("""
                            SELECT 1 FROM novedades_volvo 
                            WHERE fecha_novedad = %s AND movil = %s AND fuente_informacion = %s AND novedad = %s LIMIT 1
                        """, (fecha_nov, movil_id, fuente, obs_mec))
                    else:
                        cursor.execute("""
                            SELECT 1 FROM novedades_volvo 
                            WHERE fecha_novedad = ? AND movil = ? AND fuente_informacion = ? AND novedad = ? LIMIT 1
                        """, (fecha_nov, movil_id, fuente, obs_mec))
                        
                    if not cursor.fetchone():
                        obs_mec_clean = obs_mec.upper()
                        nov_mec_row = novedades_master.get(obs_mec_clean, {})
                        grupo_mec = nov_mec_row.get("grupo_funcion", "")
                        criticidad_mec = nov_mec_row.get("criticidad", "")
                        
                        if db_type == "MySQL":
                            cursor.execute("""
                                INSERT INTO novedades_volvo (
                                    fecha_novedad, movil, chasis, linea, contrato,
                                    fuente_informacion, novedad, observaciones,
                                    grupo_funcion, criticidad, dias, estado, fecha_correccion
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, 'PENDIENTE', NULL)
                            """, (fecha_nov, movil_id, chasis_val, linea_val, contrato_val, fuente, obs_mec, det_mec, grupo_mec, criticidad_mec))
                        else:
                            cursor.execute("""
                                INSERT INTO novedades_volvo (
                                    fecha_novedad, movil, chasis, linea, contrato,
                                    fuente_informacion, novedad, observaciones,
                                    grupo_funcion, criticidad, dias, estado, fecha_correccion
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 'PENDIENTE', NULL)
                            """, (fecha_nov, movil_id, chasis_val, linea_val, contrato_val, fuente, obs_mec, det_mec, grupo_mec, criticidad_mec))
                            
                # 4.2 Electrical Observation (only relevant for Alistamiento / Combinada checklists)
                if tipo_ins in ["ALISTAMIENTO", "COMBINADA"]:
                    obs_elec = str(r_dict['observaciones_elec']).strip() if r_dict['observaciones_elec'] else ""
                    det_elec = str(r_dict['detalles_elec']).strip() if r_dict['detalles_elec'] else ""
                    if obs_elec and obs_elec.upper() not in ["N/A", "NONE", ""]:
                        if db_type == "MySQL":
                            cursor.execute("""
                                SELECT 1 FROM novedades_volvo 
                                WHERE fecha_novedad = %s AND movil = %s AND fuente_informacion = %s AND novedad = %s LIMIT 1
                            """, (fecha_nov, movil_id, fuente, obs_elec))
                        else:
                            cursor.execute("""
                                SELECT 1 FROM novedades_volvo 
                                WHERE fecha_novedad = ? AND movil = ? AND fuente_informacion = ? AND novedad = ? LIMIT 1
                            """, (fecha_nov, movil_id, fuente, obs_elec))
                            
                        if not cursor.fetchone():
                            obs_elec_clean = obs_elec.upper()
                            nov_elec_row = novedades_master.get(obs_elec_clean, {})
                            grupo_elec = nov_elec_row.get("grupo_funcion", "")
                            criticidad_elec = nov_elec_row.get("criticidad", "")
                            
                            if db_type == "MySQL":
                                cursor.execute("""
                                    INSERT INTO novedades_volvo (
                                        fecha_novedad, movil, chasis, linea, contrato,
                                        fuente_informacion, novedad, observaciones,
                                        grupo_funcion, criticidad, dias, estado, fecha_correccion
                                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, 'PENDIENTE', NULL)
                                """, (fecha_nov, movil_id, chasis_val, linea_val, contrato_val, fuente, obs_elec, det_elec, grupo_elec, criticidad_elec))
                            else:
                                cursor.execute("""
                                    INSERT INTO novedades_volvo (
                                        fecha_novedad, movil, chasis, linea, contrato,
                                        fuente_informacion, novedad, observaciones,
                                        grupo_funcion, criticidad, dias, estado, fecha_correccion
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 'PENDIENTE', NULL)
                                """, (fecha_nov, movil_id, chasis_val, linea_val, contrato_val, fuente, obs_elec, det_elec, grupo_elec, criticidad_elec))
            conn.commit()
            print("Automatic legacy novelties migration to novedades_volvo finished successfully.")
        except Exception as ex:
            print(f"Error during legacy novelties migration: {str(ex)}")

        cursor.close()
        conn.close()

    def load_patio_data(self, patio_name, host, user, password, db_name, port):
        """
        Loads inspections for a specific patio, returning a formatted Pandas DataFrame.
        Resolves separate sub-tables through left joins, maintaining unified schema compatibility.
        """
        conn, db_type = self.connection_manager.get_connection(host, user, password, db_name, port)
        cursor = conn.cursor()
        
        query = """
            SELECT 
                i.id, i.movil,
                n.nivel_aceite_motor_si_no, n.nivel_aceite_motor_cant,
                n.nivel_refrigerante_si_no, n.nivel_refrigerante_cant,
                n.nivel_aceite_hidraulico_si_no, n.nivel_aceite_hidraulico_cant,
                n.nivel_limpiabrisas_si_no, n.nivel_limpiabrisas_cant,
                n.inspeccion_ducto_admision, n.drenar_tanques, n.observacion_niveles,
                m.revision_fugas, m.separador_humedad, m.inspeccion_embrague,
                m.inspeccion_palanca, m.inspeccion_ventilador, m.inspeccion_suspension,
                m.inspeccion_frenos, m.inspeccion_direccion,
                e.elec_luces, e.elec_tablero, e.elec_rutero, e.elec_arranque, e.elec_puertas, e.elec_baterias,
                m.detalles AS detalles, m.observaciones AS observaciones,
                e.detalles_elec AS detalles_elec, e.observaciones_elec AS observaciones_elec,
                m.insumos AS insumos,
                i.tecnico, i.fecha_registro, i.tipo_inspeccion
            FROM inspecciones i
            LEFT JOIN inspeccion_niveles n ON i.id = n.inspeccion_id
            LEFT JOIN inspeccion_mecanica m ON i.id = m.inspeccion_id
            LEFT JOIN inspeccion_electrica e ON i.id = e.inspeccion_id
            WHERE i.patio = %s AND (i.tecnico = '' OR i.tecnico IS NULL)
            ORDER BY i.movil ASC, i.id ASC
        """ if db_type == "MySQL" else """
            SELECT 
                i.id, i.movil,
                n.nivel_aceite_motor_si_no, n.nivel_aceite_motor_cant,
                n.nivel_refrigerante_si_no, n.nivel_refrigerante_cant,
                n.nivel_aceite_hidraulico_si_no, n.nivel_aceite_hidraulico_cant,
                n.nivel_limpiabrisas_si_no, n.nivel_limpiabrisas_cant,
                n.inspeccion_ducto_admision, n.drenar_tanques, n.observacion_niveles,
                m.revision_fugas, m.separador_humedad, m.inspeccion_embrague,
                m.inspeccion_palanca, m.inspeccion_ventilador, m.inspeccion_suspension,
                m.inspeccion_frenos, m.inspeccion_direccion,
                e.elec_luces, e.elec_tablero, e.elec_rutero, e.elec_arranque, e.elec_puertas, e.elec_baterias,
                m.detalles AS detalles, m.observaciones AS observaciones,
                e.detalles_elec AS detalles_elec, e.observaciones_elec AS observaciones_elec,
                m.insumos AS insumos,
                i.tecnico, i.fecha_registro, i.tipo_inspeccion
            FROM inspecciones i
            LEFT JOIN inspeccion_niveles n ON i.id = n.inspeccion_id
            LEFT JOIN inspeccion_mecanica m ON i.id = m.inspeccion_id
            LEFT JOIN inspeccion_electrica e ON i.id = e.inspeccion_id
            WHERE i.patio = ? AND (i.tecnico = '' OR i.tecnico IS NULL)
            ORDER BY i.movil ASC, i.id ASC
        """
        
        cursor.execute(query, (patio_name,))
        rows = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        df_data = pd.DataFrame(rows)
        if df_data.empty:
            df_data = pd.DataFrame(columns=UNIFIED_COLS)
        else:
            df_data.columns = UNIFIED_COLS
            
        return df_data

    def save_patio_changes_custom(self, modified_moviles, df_master, tecnico_name, patio_name, host, user, password, db_name, port):
        """
        Saves changes to the normalized database tables.
        Handles insertions of new inspection entries and linked sub-records.
        """
        db_conn, db_type = self.connection_manager.get_connection(host, user, password, db_name, port)
        db_cursor = db_conn.cursor()
        
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_count = 0
        
        for movil_id in modified_moviles:
            row_mask = df_master['Móvil'] == movil_id
            for idx in df_master[row_mask].index:
                df_master.at[idx, 'Técnico'] = tecnico_name
                df_master.at[idx, 'Fecha Registro'] = now_str
                
                row_id = df_master.at[idx, 'id']
                
                # Fluid Levels values
                lvls_data = (
                    df_master.at[idx, 'Nivel Aceite Motor (SI/NO)'], df_master.at[idx, 'Nivel Aceite Motor (L)'],
                    df_master.at[idx, 'Nivel Refrigerante (SI/NO)'], df_master.at[idx, 'Nivel Refrigerante (L)'],
                    df_master.at[idx, 'Nivel Aceite Hidráulico (SI/NO)'], df_master.at[idx, 'Nivel Aceite Hidráulico (L)'],
                    df_master.at[idx, 'Nivel Limpiabrisas (SI/NO)'], df_master.at[idx, 'Nivel Limpiabrisas (L)'],
                    df_master.at[idx, 'Inspección Ducto Admisión'], df_master.at[idx, 'Drenar Tanques'],
                    df_master.at[idx, 'Observaciones Niveles']
                )
                
                # Mechanical values
                mec_data = (
                    df_master.at[idx, 'Revisión de Fugas'], df_master.at[idx, 'Separador de Humedad'],
                    df_master.at[idx, 'Inspección Embrague'], df_master.at[idx, 'Inspección Palanca'],
                    df_master.at[idx, 'Inspección Ventilador'], df_master.at[idx, 'Inspección Suspensión'],
                    df_master.at[idx, 'Inspección Frenos'], df_master.at[idx, 'Inspección Dirección'],
                    df_master.at[idx, 'Detalles / Novedades'], df_master.at[idx, 'Observaciones'],
                    df_master.at[idx, 'Insumos / SAP']
                )
                
                # Electrical values
                elec_data = (
                    df_master.at[idx, 'Elec. Luces'], df_master.at[idx, 'Elec. Tablero'],
                    df_master.at[idx, 'Elec. Rutero'], df_master.at[idx, 'Elec. Arranque'],
                    df_master.at[idx, 'Elec. Puertas'], df_master.at[idx, 'Elec. Baterías'],
                    df_master.at[idx, 'Detalles Eléctrico'], df_master.at[idx, 'Observaciones Eléctrico']
                )

                tipo_novedad_val = df_master.at[idx, 'Tipo Novedad']
                if not tipo_novedad_val:
                    tipo_novedad_val = 'ALISTAMIENTO'

                if row_id is not None and pd.notnull(row_id) and str(row_id).strip() != "":
                    # 1. Update parent inspection entry
                    inspeccion_id = int(row_id)
                    if db_type == "MySQL":
                        db_cursor.execute("""
                            UPDATE inspecciones SET tecnico = %s, fecha_registro = %s, tipo_inspeccion = %s WHERE id = %s
                        """, (tecnico_name, now_str, tipo_novedad_val, inspeccion_id))
                    else:
                        db_cursor.execute("""
                            UPDATE inspecciones SET tecnico = ?, fecha_registro = ?, tipo_inspeccion = ? WHERE id = ?
                        """, (tecnico_name, now_str, tipo_novedad_val, inspeccion_id))

                    # 2. Update/Insert child levels record
                    if db_type == "MySQL":
                        db_cursor.execute("""
                            INSERT INTO inspeccion_niveles (
                                inspeccion_id, nivel_aceite_motor_si_no, nivel_aceite_motor_cant,
                                nivel_refrigerante_si_no, nivel_refrigerante_cant,
                                nivel_aceite_hidraulico_si_no, nivel_aceite_hidraulico_cant,
                                nivel_limpiabrisas_si_no, nivel_limpiabrisas_cant,
                                inspeccion_ducto_admision, drenar_tanques, observacion_niveles
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE 
                                nivel_aceite_motor_si_no=VALUES(nivel_aceite_motor_si_no), nivel_aceite_motor_cant=VALUES(nivel_aceite_motor_cant),
                                nivel_refrigerante_si_no=VALUES(nivel_refrigerante_si_no), nivel_refrigerante_cant=VALUES(nivel_refrigerante_cant),
                                nivel_aceite_hidraulico_si_no=VALUES(nivel_aceite_hidraulico_si_no), nivel_aceite_hidraulico_cant=VALUES(nivel_aceite_hidraulico_cant),
                                nivel_limpiabrisas_si_no=VALUES(nivel_limpiabrisas_si_no), nivel_limpiabrisas_cant=VALUES(nivel_limpiabrisas_cant),
                                inspeccion_ducto_admision=VALUES(inspeccion_ducto_admision), drenar_tanques=VALUES(drenar_tanques),
                                observacion_niveles=VALUES(observacion_niveles)
                        """, (inspeccion_id,) + lvls_data)
                    else:
                        # SQLite Upsert
                        db_cursor.execute("""
                            INSERT INTO inspeccion_niveles (
                                inspeccion_id, nivel_aceite_motor_si_no, nivel_aceite_motor_cant,
                                nivel_refrigerante_si_no, nivel_refrigerante_cant,
                                nivel_aceite_hidraulico_si_no, nivel_aceite_hidraulico_cant,
                                nivel_limpiabrisas_si_no, nivel_limpiabrisas_cant,
                                inspeccion_ducto_admision, drenar_tanques, observacion_niveles
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ON CONFLICT(inspeccion_id) DO UPDATE SET 
                                nivel_aceite_motor_si_no=excluded.nivel_aceite_motor_si_no, nivel_aceite_motor_cant=excluded.nivel_aceite_motor_cant,
                                nivel_refrigerante_si_no=excluded.nivel_refrigerante_si_no, nivel_refrigerante_cant=excluded.nivel_refrigerante_cant,
                                nivel_aceite_hidraulico_si_no=excluded.nivel_aceite_hidraulico_si_no, nivel_aceite_hidraulico_cant=excluded.nivel_aceite_hidraulico_cant,
                                nivel_limpiabrisas_si_no=excluded.nivel_limpiabrisas_si_no, nivel_limpiabrisas_cant=excluded.nivel_limpiabrisas_cant,
                                inspeccion_ducto_admision=excluded.inspeccion_ducto_admision, drenar_tanques=excluded.drenar_tanques,
                                observacion_niveles=excluded.observacion_niveles
                        """, (inspeccion_id,) + lvls_data)

                    # 3. Update/Insert child mechanical record (Upsert based on parent inspeccion_id)
                    # We check if a row exists in inspeccion_mecanica for this parent
                    if db_type == "MySQL":
                        db_cursor.execute("SELECT id FROM inspeccion_mecanica WHERE inspeccion_id=%s LIMIT 1", (inspeccion_id,))
                    else:
                        db_cursor.execute("SELECT id FROM inspeccion_mecanica WHERE inspeccion_id=? LIMIT 1", (inspeccion_id,))
                    mec_row = db_cursor.fetchone()
                    
                    if mec_row:
                        mec_child_id = mec_row[0] if not isinstance(mec_row, dict) else mec_row['id']
                        if db_type == "MySQL":
                            db_cursor.execute("""
                                UPDATE inspeccion_mecanica SET
                                    revision_fugas=%s, separador_humedad=%s, inspeccion_embrague=%s,
                                    inspeccion_palanca=%s, inspeccion_ventilador=%s, inspeccion_suspension=%s,
                                    inspeccion_frenos=%s, inspeccion_direccion=%s, detalles=%s, observaciones=%s, insumos=%s
                                WHERE id=%s
                            """, mec_data + (mec_child_id,))
                        else:
                            db_cursor.execute("""
                                UPDATE inspeccion_mecanica SET
                                    revision_fugas=?, separador_humedad=?, inspeccion_embrague=?,
                                    inspeccion_palanca=?, inspeccion_ventilador=?, inspeccion_suspension=?,
                                    inspeccion_frenos=?, inspeccion_direccion=?, detalles=?, observaciones=?, insumos=?
                                WHERE id=?
                            """, mec_data + (mec_child_id,))
                    else:
                        if db_type == "MySQL":
                            db_cursor.execute("""
                                INSERT INTO inspeccion_mecanica (
                                    inspeccion_id, revision_fugas, separador_humedad, inspeccion_embrague,
                                    inspeccion_palanca, inspeccion_ventilador, inspeccion_suspension,
                                    inspeccion_frenos, inspeccion_direccion, detalles, observaciones, insumos
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, (inspeccion_id,) + mec_data)
                        else:
                            db_cursor.execute("""
                                INSERT INTO inspeccion_mecanica (
                                    inspeccion_id, revision_fugas, separador_humedad, inspeccion_embrague,
                                    inspeccion_palanca, inspeccion_ventilador, inspeccion_suspension,
                                    inspeccion_frenos, inspeccion_direccion, detalles, observaciones, insumos
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (inspeccion_id,) + mec_data)

                    # 4. Update/Insert child electrical record (Upsert based on parent inspeccion_id)
                    if db_type == "MySQL":
                        db_cursor.execute("SELECT id FROM inspeccion_electrica WHERE inspeccion_id=%s LIMIT 1", (inspeccion_id,))
                    else:
                        db_cursor.execute("SELECT id FROM inspeccion_electrica WHERE inspeccion_id=? LIMIT 1", (inspeccion_id,))
                    elec_row = db_cursor.fetchone()
                    
                    if elec_row:
                        elec_child_id = elec_row[0] if not isinstance(elec_row, dict) else elec_row['id']
                        if db_type == "MySQL":
                            db_cursor.execute("""
                                UPDATE inspeccion_electrica SET
                                    elec_luces=%s, elec_tablero=%s, elec_rutero=%s, elec_arranque=%s,
                                    elec_puertas=%s, elec_baterias=%s, detalles_elec=%s, observaciones_elec=%s
                                WHERE id=%s
                            """, elec_data + (elec_child_id,))
                        else:
                            db_cursor.execute("""
                                UPDATE inspeccion_electrica SET
                                    elec_luces=?, elec_tablero=?, elec_rutero=?, elec_arranque=?,
                                    elec_puertas=?, elec_baterias=?, detalles_elec=?, observaciones_elec=?
                                WHERE id=?
                            """, elec_data + (elec_child_id,))
                    else:
                        if db_type == "MySQL":
                            db_cursor.execute("""
                                INSERT INTO inspeccion_electrica (
                                    inspeccion_id, elec_luces, elec_tablero, elec_rutero,
                                    elec_arranque, elec_puertas, elec_baterias, detalles_elec, observaciones_elec
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, (inspeccion_id,) + elec_data)
                        else:
                            db_cursor.execute("""
                                INSERT INTO inspeccion_electrica (
                                    inspeccion_id, elec_luces, elec_tablero, elec_rutero,
                                    elec_arranque, elec_puertas, elec_baterias, detalles_elec, observaciones_elec
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (inspeccion_id,) + elec_data)

                else:
                    # 1. Insert new parent inspection
                    if db_type == "MySQL":
                        db_cursor.execute("""
                            INSERT INTO inspecciones (movil, patio, tecnico, fecha_registro, tipo_inspeccion)
                            VALUES (%s, %s, %s, %s, %s)
                        """, (movil_id, patio_name, tecnico_name, now_str, tipo_novedad_val))
                        inspeccion_id = db_cursor.lastrowid
                    else:
                        db_cursor.execute("""
                            INSERT INTO inspecciones (movil, patio, tecnico, fecha_registro, tipo_inspeccion)
                            VALUES (?, ?, ?, ?, ?)
                        """, (movil_id, patio_name, tecnico_name, now_str, tipo_novedad_val))
                        inspeccion_id = db_cursor.lastrowid
                    
                    df_master.at[idx, 'id'] = inspeccion_id

                    # 2. Insert levels sub-record
                    if db_type == "MySQL":
                        db_cursor.execute("""
                            INSERT INTO inspeccion_niveles (
                                inspeccion_id, nivel_aceite_motor_si_no, nivel_aceite_motor_cant,
                                nivel_refrigerante_si_no, nivel_refrigerante_cant,
                                nivel_aceite_hidraulico_si_no, nivel_aceite_hidraulico_cant,
                                nivel_limpiabrisas_si_no, nivel_limpiabrisas_cant,
                                inspeccion_ducto_admision, drenar_tanques, observacion_niveles
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (inspeccion_id,) + lvls_data)
                    else:
                        db_cursor.execute("""
                            INSERT INTO inspeccion_niveles (
                                inspeccion_id, nivel_aceite_motor_si_no, nivel_aceite_motor_cant,
                                nivel_refrigerante_si_no, nivel_refrigerante_cant,
                                nivel_aceite_hidraulico_si_no, nivel_aceite_hidraulico_cant,
                                nivel_limpiabrisas_si_no, nivel_limpiabrisas_cant,
                                inspeccion_ducto_admision, drenar_tanques, observacion_niveles
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (inspeccion_id,) + lvls_data)

                    # 3. Insert mechanical sub-record
                    if db_type == "MySQL":
                        db_cursor.execute("""
                            INSERT INTO inspeccion_mecanica (
                                inspeccion_id, revision_fugas, separador_humedad, inspeccion_embrague,
                                inspeccion_palanca, inspeccion_ventilador, inspeccion_suspension,
                                inspeccion_frenos, inspeccion_direccion, detalles, observaciones, insumos
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (inspeccion_id,) + mec_data)
                    else:
                        db_cursor.execute("""
                            INSERT INTO inspeccion_mecanica (
                                inspeccion_id, revision_fugas, separador_humedad, inspeccion_embrague,
                                inspeccion_palanca, inspeccion_ventilador, inspeccion_suspension,
                                inspeccion_frenos, inspeccion_direccion, detalles, observaciones, insumos
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (inspeccion_id,) + mec_data)

                    # 4. Insert electrical sub-record
                    if db_type == "MySQL":
                        db_cursor.execute("""
                            INSERT INTO inspeccion_electrica (
                                inspeccion_id, elec_luces, elec_tablero, elec_rutero,
                                elec_arranque, elec_puertas, elec_baterias, detalles_elec, observaciones_elec
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (inspeccion_id,) + elec_data)
                    else:
                        db_cursor.execute("""
                            INSERT INTO inspeccion_electrica (
                                inspeccion_id, elec_luces, elec_tablero, elec_rutero,
                                elec_arranque, elec_puertas, elec_baterias, detalles_elec, observaciones_elec
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (inspeccion_id,) + elec_data)

                updated_count += 1
                
        try:
            from src.services.export_service import ExportService
            fleet_info = ExportService.load_fleet_master()
            novedades_master = ExportService.load_novedades_master()
            
            # Extract and insert novelties
            novedades_a_guardar = []
            for movil_id in modified_moviles:
                movil_clean = str(movil_id).strip()
                fleet_row = fleet_info.get(movil_clean, {})
                chasis_val = fleet_row.get("chasis", "")
                linea_val = fleet_row.get("linea", "")
                contrato_val = fleet_row.get("contrato", "")
                
                row_mask = df_master['Móvil'] == movil_id
                for idx in df_master[row_mask].index:
                    tipo_nov_val = df_master.at[idx, 'Tipo Novedad']
                    
                    if tipo_nov_val == "ALISTAMIENTO":
                        # 1. Mechanical novelty
                        obs_mecanica = str(df_master.at[idx, 'Observaciones']).strip()
                        if obs_mecanica and obs_mecanica.upper() != "N/A":
                            det_mecanico = str(df_master.at[idx, 'Detalles / Novedades']).strip()
                            obs_mec_clean = obs_mecanica.upper()
                            nov_mec_row = novedades_master.get(obs_mec_clean, {})
                            novedades_a_guardar.append({
                                "fecha_novedad": now_str[:10],
                                "movil": movil_id,
                                "chasis": chasis_val,
                                "linea": linea_val,
                                "contrato": contrato_val,
                                "fuente_informacion": "ALISTAMIENTO",
                                "novedad": obs_mecanica,
                                "observaciones": det_mecanico,
                                "grupo_funcion": nov_mec_row.get("grupo_funcion", ""),
                                "criticidad": nov_mec_row.get("criticidad", ""),
                                "dias": 0,
                                "estado": "PENDIENTE",
                                "fecha_correccion": None
                            })
                        
                        # 2. Electrical novelty
                        obs_electrica = str(df_master.at[idx, 'Observaciones Eléctrico']).strip()
                        if obs_electrica and obs_electrica.upper() != "N/A":
                            det_electrico = str(df_master.at[idx, 'Detalles Eléctrico']).strip()
                            obs_elec_clean = obs_electrica.upper()
                            nov_elec_row = novedades_master.get(obs_elec_clean, {})
                            novedades_a_guardar.append({
                                "fecha_novedad": now_str[:10],
                                "movil": movil_id,
                                "chasis": chasis_val,
                                "linea": linea_val,
                                "contrato": contrato_val,
                                "fuente_informacion": "ALISTAMIENTO",
                                "novedad": obs_electrica,
                                "observaciones": det_electrico,
                                "grupo_funcion": nov_elec_row.get("grupo_funcion", ""),
                                "criticidad": nov_elec_row.get("criticidad", ""),
                                "dias": 0,
                                "estado": "PENDIENTE",
                                "fecha_correccion": None
                            })
                    
                    elif tipo_nov_val == "CORRECTIVO":
                        obs_mecanica = str(df_master.at[idx, 'Observaciones']).strip()
                        if obs_mecanica:
                            det_mecanico = str(df_master.at[idx, 'Detalles / Novedades']).strip()
                            obs_mec_clean = obs_mecanica.upper()
                            nov_mec_row = novedades_master.get(obs_mec_clean, {})
                            novedades_a_guardar.append({
                                "fecha_novedad": now_str[:10],
                                "movil": movil_id,
                                "chasis": chasis_val,
                                "linea": linea_val,
                                "contrato": contrato_val,
                                "fuente_informacion": "MANTENIMIENTO CORRECTIVO",
                                "novedad": obs_mecanica,
                                "observaciones": det_mecanico,
                                "grupo_funcion": nov_mec_row.get("grupo_funcion", ""),
                                "criticidad": nov_mec_row.get("criticidad", ""),
                                "dias": 0,
                                "estado": "PENDIENTE",
                                "fecha_correccion": None
                            })
                            
                    elif tipo_nov_val == "PREVENTIVO":
                        obs_mecanica = str(df_master.at[idx, 'Observaciones']).strip()
                        if obs_mecanica:
                            det_mecanico = str(df_master.at[idx, 'Detalles / Novedades']).strip()
                            obs_mec_clean = obs_mecanica.upper()
                            nov_mec_row = novedades_master.get(obs_mec_clean, {})
                            novedades_a_guardar.append({
                                "fecha_novedad": now_str[:10],
                                "movil": movil_id,
                                "chasis": chasis_val,
                                "linea": linea_val,
                                "contrato": contrato_val,
                                "fuente_informacion": "MANTENIMIENTO PREVENTIVO",
                                "novedad": obs_mecanica,
                                "observaciones": det_mecanico,
                                "grupo_funcion": nov_mec_row.get("grupo_funcion", ""),
                                "criticidad": nov_mec_row.get("criticidad", ""),
                                "dias": 0,
                                "estado": "PENDIENTE",
                                "fecha_correccion": None
                            })
            
            # Save to DB
            for nov in novedades_a_guardar:
                if db_type == "MySQL":
                    db_cursor.execute("""
                        INSERT INTO novedades_volvo (
                            fecha_novedad, movil, chasis, linea, contrato,
                            fuente_informacion, novedad, observaciones,
                            grupo_funcion, criticidad, dias, estado, fecha_correccion
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        nov["fecha_novedad"], nov["movil"], nov["chasis"], nov["linea"], nov["contrato"],
                        nov["fuente_informacion"], nov["novedad"], nov["observaciones"],
                        nov["grupo_funcion"], nov["criticidad"], nov["dias"], nov["estado"], nov["fecha_correccion"]
                    ))
                else:
                    db_cursor.execute("""
                        INSERT INTO novedades_volvo (
                            fecha_novedad, movil, chasis, linea, contrato,
                            fuente_informacion, novedad, observaciones,
                            grupo_funcion, criticidad, dias, estado, fecha_correccion
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        nov["fecha_novedad"], nov["movil"], nov["chasis"], nov["linea"], nov["contrato"],
                        nov["fuente_informacion"], nov["novedad"], nov["observaciones"],
                        nov["grupo_funcion"], nov["criticidad"], nov["dias"], nov["estado"], nov["fecha_correccion"]
                    ))

            # Save to Excel
            if novedades_a_guardar:
                import os
                import openpyxl
                from openpyxl.styles import Border, Side, Font, Alignment
                
                excel_path = os.path.join(os.getcwd(), "Formatos_Alistamiento", "FORMATO NOVEDADES VOLVO.xlsx")
                if not os.path.exists(excel_path):
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Novedades"
                else:
                    try:
                        wb = openpyxl.load_workbook(excel_path)
                        ws = wb.active
                    except PermissionError:
                        raise PermissionError(
                            "No se pudo guardar: El archivo 'FORMATO NOVEDADES VOLVO.xlsx' está abierto en Excel. "
                            "Por favor, ciérrelo en Excel e intente guardar de nuevo."
                        )
                    except Exception as ex:
                        raise Exception(f"No se pudo cargar el archivo Excel de novedades: {str(ex)}")
                
                expected_headers = [
                    "FECHA NOVEDAD", "ID", "CHASIS", "LINEA", "CONTRATO", 
                    "FUENTE DE INFORMACION", "NOVEDAD", "OBSERVACIONES", 
                    "GRUPO FUNCION", "CRITICIDAD", "DIAS", "ESTADO", "FECHA CORRECCION",
                    "TECNICO QUE HACE CORRECCION", "INSUMOS / PARTE NUMEROS", "CANTIDAD"
                ]
                
                # Check headers (at row 5 in the template)
                headers = [cell.value for cell in ws[5]] if ws.max_row >= 5 else []
                headers_clean = [str(h).strip().upper() if h is not None else "" for h in headers]
                
                if not any(headers_clean):
                    for col_num, h_name in enumerate(expected_headers, 1):
                        ws.cell(row=5, column=col_num, value=h_name)
                    headers_clean = [h.upper() for h in expected_headers]
                
                col_map = {}
                for col_idx, h_clean in enumerate(headers_clean, 1):
                    if h_clean:
                        col_map[h_clean] = col_idx
                
                for h_name in expected_headers:
                    h_upper = h_name.upper()
                    if h_upper not in col_map:
                        next_col = len(col_map) + 1
                        ws.cell(row=5, column=next_col, value=h_name)
                        col_map[h_upper] = next_col
                
                thin_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                font_body = Font(name='Arial', size=11)
                align_center = Alignment(horizontal='center', vertical='center')
                align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                from openpyxl.utils import get_column_letter
                col_fecha_nov_letter = get_column_letter(col_map["FECHA NOVEDAD"])
                col_fecha_corr_letter = get_column_letter(col_map["FECHA CORRECCION"])

                for nov in novedades_a_guardar:
                    next_row = ws.max_row + 1
                    
                    f_novedad_val = nov["fecha_novedad"]
                    if len(f_novedad_val) > 10:
                        f_novedad_val = f_novedad_val[:10]
                        
                    ws.cell(row=next_row, column=col_map["FECHA NOVEDAD"], value=f_novedad_val)
                    ws.cell(row=next_row, column=col_map["ID"], value=nov["movil"])
                    ws.cell(row=next_row, column=col_map["CHASIS"], value=nov["chasis"])
                    ws.cell(row=next_row, column=col_map["LINEA"], value=nov["linea"])
                    ws.cell(row=next_row, column=col_map["CONTRATO"], value=nov["contrato"])
                    ws.cell(row=next_row, column=col_map["FUENTE DE INFORMACION"], value=nov["fuente_informacion"])
                    ws.cell(row=next_row, column=col_map["NOVEDAD"], value=nov["novedad"])
                    ws.cell(row=next_row, column=col_map["OBSERVACIONES"], value=nov["observaciones"])
                    ws.cell(row=next_row, column=col_map["GRUPO FUNCION"], value=nov["grupo_funcion"])
                    ws.cell(row=next_row, column=col_map["CRITICIDAD"], value=nov["criticidad"])
                    ws.cell(row=next_row, column=col_map["ESTADO"], value=nov["estado"])
                    ws.cell(row=next_row, column=col_map["FECHA CORRECCION"], value=nov["fecha_correccion"])
                    ws.cell(row=next_row, column=col_map["TECNICO QUE HACE CORRECCION"], value=nov.get("tecnico_correccion", ""))
                    ws.cell(row=next_row, column=col_map["INSUMOS / PARTE NUMEROS"], value=nov.get("insumos_usados", ""))
                    ws.cell(row=next_row, column=col_map["CANTIDAD"], value=nov.get("cantidad", ""))
                    
                    for c_idx in col_map.values():
                        cell = ws.cell(row=next_row, column=c_idx)
                        cell.border = thin_border
                        cell.font = font_body
                        if c_idx in [col_map["FECHA NOVEDAD"], col_map["ID"], col_map["DIAS"], col_map["ESTADO"], col_map["FECHA CORRECCION"]]:
                            cell.alignment = align_center
                        else:
                            cell.alignment = align_left

                # Formulate the DIAS column for ALL rows (from row 6 to ws.max_row)
                col_estado_letter = get_column_letter(col_map["ESTADO"])
                for r_idx in range(6, ws.max_row + 1):
                    dias_formula = f'=IF({col_estado_letter}{r_idx}="CORREGIDA", 0, TODAY()-{col_fecha_nov_letter}{r_idx})'
                    ws.cell(row=r_idx, column=col_map["DIAS"], value=dias_formula)
                
                try:
                    wb.save(excel_path)
                except PermissionError:
                    raise PermissionError(
                        "No se pudo guardar: El archivo 'FORMATO NOVEDADES VOLVO.xlsx' está abierto en Excel. "
                        "Por favor, ciérrelo en Excel e intente guardar de nuevo."
                    )
                except Exception as ex:
                    raise Exception(f"No se pudo guardar el archivo Excel de novedades: {str(ex)}")

            db_conn.commit()
        except Exception as e:
            db_conn.rollback()
            raise e
        finally:
            db_cursor.close()
            db_conn.close()
            
        return updated_count

    def load_all_data(self, host, user, password, db_name, port):
        """
        Loads all completed inspections for reporting, denormalized across sub-tables.
        """
        conn, db_type = self.connection_manager.get_connection(host, user, password, db_name, port)
        cursor = conn.cursor()
        
        query = """
            SELECT 
                i.id, i.patio, i.movil,
                n.nivel_aceite_motor_si_no, n.nivel_aceite_motor_cant,
                n.nivel_refrigerante_si_no, n.nivel_refrigerante_cant,
                n.nivel_aceite_hidraulico_si_no, n.nivel_aceite_hidraulico_cant,
                n.nivel_limpiabrisas_si_no, n.nivel_limpiabrisas_cant,
                n.inspeccion_ducto_admision, n.drenar_tanques, n.observacion_niveles,
                m.revision_fugas, m.separador_humedad, m.inspeccion_embrague,
                m.inspeccion_palanca, m.inspeccion_ventilador, m.inspeccion_suspension,
                m.inspeccion_frenos, m.inspeccion_direccion,
                e.elec_luces, e.elec_tablero, e.elec_rutero, e.elec_arranque, e.elec_puertas, e.elec_baterias,
                m.detalles AS detalles, m.observaciones AS observaciones,
                e.detalles_elec AS detalles_elec, e.observaciones_elec AS observaciones_elec,
                m.insumos AS insumos,
                i.tecnico, i.fecha_registro, i.tipo_inspeccion
            FROM inspecciones i
            LEFT JOIN inspeccion_niveles n ON i.id = n.inspeccion_id
            LEFT JOIN inspeccion_mecanica m ON i.id = m.inspeccion_id
            LEFT JOIN inspeccion_electrica e ON i.id = e.inspeccion_id
            WHERE i.tecnico != '' AND i.tecnico IS NOT NULL
            ORDER BY i.patio ASC, i.movil ASC, i.id ASC
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        df_data = pd.DataFrame(rows)
        cols_all = ['id', 'Patio'] + UNIFIED_COLS[1:]
        if df_data.empty:
            df_data = pd.DataFrame(columns=cols_all)
        else:
            df_data.columns = cols_all
            
        return df_data

    def delete_record_by_id(self, record_id, host, user, password, db_name, port):
        """
        Deletes a specific parent inspection record, cascade-deleting linked child rows.
        """
        conn, db_type = self.connection_manager.get_connection(host, user, password, db_name, port)
        cursor = conn.cursor()
        
        if db_type == "MySQL":
            cursor.execute("DELETE FROM inspecciones WHERE id = %s", (record_id,))
        else:
            cursor.execute("DELETE FROM inspecciones WHERE id = ?", (record_id,))
            
        conn.commit()
        cursor.close()
        conn.close()

    def load_novedades_volvo(self, host, user, password, db_name, port):
        """
        Loads all records from the novedades_volvo table.
        """
        conn, db_type = self.connection_manager.get_connection(host, user, password, db_name, port)
        cursor = conn.cursor()
        
        query = """
            SELECT fecha_novedad, movil, chasis, linea, contrato, 
                   fuente_informacion, novedad, observaciones, 
                   grupo_funcion, criticidad, dias, estado, fecha_correccion,
                   tecnico_correccion, insumos_usados, cantidad
            FROM novedades_volvo 
            ORDER BY id ASC
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        df = pd.DataFrame(rows)
        cols = [
            "FECHA NOVEDAD", "ID", "CHASIS", "LINEA", "CONTRATO", 
            "FUENTE DE INFORMACION", "NOVEDAD", "OBSERVACIONES", 
            "GRUPO FUNCION", "CRITICIDAD", "DIAS", "ESTADO", "FECHA CORRECCION",
            "TECNICO QUE HACE CORRECCION", "INSUMOS / PARTE NUMEROS", "CANTIDAD"
        ]
        if df.empty:
            df = pd.DataFrame(columns=cols)
        else:
            df.columns = cols
            try:
                import datetime
                today = datetime.date.today()
                
                def get_days_open(row):
                    try:
                        f_nov_str = str(row["FECHA NOVEDAD"]).strip().split()[0]
                        dt_nov = datetime.datetime.strptime(f_nov_str, "%Y-%m-%d").date()
                    except Exception:
                        return 0
                    
                    estado = str(row["ESTADO"]).strip().upper()
                    if estado == "CORREGIDA":
                        return 0
                    
                    delta = (today - dt_nov).days
                    return max(0, delta)
                
                df["DIAS"] = df.apply(get_days_open, axis=1)
            except Exception as ex:
                print(f"Error calculating dynamic DIAS: {str(ex)}")
        return df

    def load_novedades_volvo_raw(self, host, user, password, db_name, port):
        """
        Loads all records from the novedades_volvo table, including the primary key ID and correction details.
        """
        conn, db_type = self.connection_manager.get_connection(host, user, password, db_name, port)
        cursor = conn.cursor()
        
        query = """
            SELECT id, fecha_novedad, movil, chasis, linea, contrato, 
                   fuente_informacion, novedad, observaciones, 
                   grupo_funcion, criticidad, dias, estado, fecha_correccion,
                   tecnico_correccion, insumos_usados, cantidad
            FROM novedades_volvo 
            ORDER BY id ASC
        """
        try:
            cursor.execute(query)
            rows = cursor.fetchall()
        except Exception:
            # Fallback if columns don't exist yet (before init_db runs)
            query_fallback = """
                SELECT id, fecha_novedad, movil, chasis, linea, contrato, 
                       fuente_informacion, novedad, observaciones, 
                       grupo_funcion, criticidad, dias, estado, fecha_correccion
                FROM novedades_volvo 
                ORDER BY id ASC
            """
            cursor.execute(query_fallback)
            rows = cursor.fetchall()
            
        cursor.close()
        conn.close()
        
        df = pd.DataFrame(rows)
        if df.empty:
            cols = [
                "id", "fecha_novedad", "movil", "chasis", "linea", "contrato",
                "fuente_informacion", "novedad", "observaciones", 
                "grupo_funcion", "criticidad", "dias", "estado", "fecha_correccion",
                "tecnico_correccion", "insumos_usados", "cantidad"
            ]
            df = pd.DataFrame(columns=cols)
        else:
            # Make sure all required columns are in the DataFrame even if fallback was used
            if "tecnico_correccion" not in df.columns:
                df["tecnico_correccion"] = ""
            if "insumos_usados" not in df.columns:
                df["insumos_usados"] = ""
            if "cantidad" not in df.columns:
                df["cantidad"] = ""
                
            try:
                import datetime
                today = datetime.date.today()
                
                def get_days_open_raw(row):
                    try:
                        f_nov_str = str(row["fecha_novedad"]).strip().split()[0]
                        dt_nov = datetime.datetime.strptime(f_nov_str, "%Y-%m-%d").date()
                    except Exception:
                        return 0
                    
                    estado = str(row["estado"]).strip().upper()
                    if estado == "CORREGIDA":
                        return 0
                    
                    delta = (today - dt_nov).days
                    return max(0, delta)
                
                df["dias"] = df.apply(get_days_open_raw, axis=1)
            except Exception as ex:
                print(f"Error calculating dynamic dias: {str(ex)}")
        return df

    def resolver_novedad(self, host, user, password, db_name, port, db_id, tecnico_correccion, insumos_usados, cantidad, fecha_correccion):
        """
        Marks a novelty as resolved (CORREGIDA) with the technician name, used parts/insumos, and quantity.
        Synchronizes this update in both the active database and the Excel file if it exists.
        """
        import os
        conn, db_type = self.connection_manager.get_connection(host, user, password, db_name, port)
        cursor = conn.cursor()
        
        # 1. Retrieve the novelty details to find matching row in Excel
        if db_type == "MySQL":
            cursor.execute("""
                SELECT movil, novedad, fecha_novedad FROM novedades_volvo WHERE id = %s
            """, (db_id,))
        else:
            cursor.execute("""
                SELECT movil, novedad, fecha_novedad FROM novedades_volvo WHERE id = ?
            """, (db_id,))
        row_info = cursor.fetchone()
        
        # 2. Update database record
        if db_type == "MySQL":
            cursor.execute("""
                UPDATE novedades_volvo 
                SET estado = 'CORREGIDA', 
                    fecha_correccion = %s,
                    tecnico_correccion = %s,
                    insumos_usados = %s,
                    cantidad = %s
                WHERE id = %s
            """, (fecha_correccion, tecnico_correccion, insumos_usados, cantidad, db_id))
        else:
            cursor.execute("""
                UPDATE novedades_volvo 
                SET estado = 'CORREGIDA', 
                    fecha_correccion = ?,
                    tecnico_correccion = ?,
                    insumos_usados = ?,
                    cantidad = ?
                WHERE id = ?
            """, (fecha_correccion, tecnico_correccion, insumos_usados, cantidad, db_id))
            
        conn.commit()
        cursor.close()
        conn.close()
        
        # 3. Update static Excel file if it exists
        if row_info:
            target_movil = str(row_info["movil"]).strip()
            target_novedad = str(row_info["novedad"]).strip()
            target_fecha = str(row_info["fecha_novedad"])[:10]
            
            excel_path = os.path.join(os.getcwd(), "Formatos_Alistamiento", "FORMATO NOVEDADES VOLVO.xlsx")
            if os.path.exists(excel_path):
                import openpyxl
                try:
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active
                    
                    # Check headers (row 5)
                    headers = [cell.value for cell in ws[5]] if ws.max_row >= 5 else []
                    headers_clean = [str(h).strip().upper() if h is not None else "" for h in headers]
                    col_map = {h: idx for idx, h in enumerate(headers_clean, 1) if h}
                    
                    if "ID" in col_map and "NOVEDAD" in col_map and "FECHA NOVEDAD" in col_map:
                        id_col = col_map["ID"]
                        nov_col = col_map["NOVEDAD"]
                        fecha_col = col_map["FECHA NOVEDAD"]
                        estado_col = col_map.get("ESTADO")
                        fecha_corr_col = col_map.get("FECHA CORRECCION")
                        tecnico_corr_col = col_map.get("TECNICO QUE HACE CORRECCION")
                        insumos_corr_col = col_map.get("INSUMOS / PARTE NUMEROS")
                        cantidad_corr_col = col_map.get("CANTIDAD")
                        
                        from openpyxl.utils import get_column_letter
                        col_fecha_nov_letter = get_column_letter(fecha_col)
                        col_fecha_corr_letter = get_column_letter(fecha_corr_col) if fecha_corr_col else 'M'
                        
                        for r_idx in range(6, ws.max_row + 1):
                            row_movil = str(ws.cell(row=r_idx, column=id_col).value).strip()
                            row_novedad = str(ws.cell(row=r_idx, column=nov_col).value).strip()
                            
                            row_fecha_val = ws.cell(row=r_idx, column=fecha_col).value
                            row_fecha = str(row_fecha_val)[:10] if row_fecha_val else ""
                            
                            if row_movil == target_movil and row_novedad == target_novedad and row_fecha == target_fecha:
                                if estado_col:
                                    ws.cell(row=r_idx, column=estado_col, value="CORREGIDA")
                                if fecha_corr_col:
                                    ws.cell(row=r_idx, column=fecha_corr_col, value=fecha_correccion)
                                # Optionally set technician and insumos if headers exist
                                if tecnico_corr_col:
                                    ws.cell(row=r_idx, column=tecnico_corr_col, value=tecnico_correccion)
                                if insumos_corr_col:
                                    ws.cell(row=r_idx, column=insumos_corr_col, value=insumos_usados)
                                if cantidad_corr_col:
                                    ws.cell(row=r_idx, column=cantidad_corr_col, value=cantidad)
                                break
                                
                        # Formulate the DIAS column for ALL rows (from row 6 to ws.max_row)
                        if "DIAS" in col_map:
                            dias_col = col_map["DIAS"]
                            col_estado_letter = get_column_letter(estado_col) if estado_col else 'L'
                            for r_idx in range(6, ws.max_row + 1):
                                dias_formula = f'=IF({col_estado_letter}{r_idx}="CORREGIDA", 0, TODAY()-{col_fecha_nov_letter}{r_idx})'
                                ws.cell(row=r_idx, column=dias_col, value=dias_formula)
                    wb.save(excel_path)
                except PermissionError:
                    raise PermissionError(
                        "No se pudo guardar: El archivo 'FORMATO NOVEDADES VOLVO.xlsx' está abierto en Excel. "
                        "Por favor, ciérrelo en Excel e intente de nuevo."
                    )
                except Exception as ex:
                    # Log or ignore minor excel sync errors so the DB write isn't rolled back
                    print(f"Warning: could not update Excel sheet: {str(ex)}")
