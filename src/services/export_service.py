# src/services/export_service.py
import io
import os
import datetime
import pandas as pd
import streamlit as st

class ExportService:
    @staticmethod
    @st.cache_data(show_spinner=False)
    def load_fleet_master():
        """
        Loads the fleet master Excel file (Flota maestra.xlsx) using a powershell
        subprocess copy to avoid exclusive Windows file locking errors.
        Returns a dictionary mapping mobile ID -> {'chasis': ..., 'linea': ..., 'contrato': ...}
        """
        import openpyxl
        import subprocess
        
        flota_original = os.path.join(os.getcwd(), "Flota Maestra", "Flota maestra.xlsx")
        flota_temp = os.path.join(os.getcwd(), "scratch", "temp_flota_load.xlsx")
        
        fleet_dict = {}
        if not os.path.exists(flota_original):
            return fleet_dict
            
        try:
            os.makedirs(os.path.dirname(flota_temp), exist_ok=True)
            subprocess.run(
                ["powershell", "-Command", f"Copy-Item '{flota_original}' '{flota_temp}'"],
                check=True,
                capture_output=True
            )
            
            wb = openpyxl.load_workbook(flota_temp, read_only=True, data_only=True)
            ws = wb.active
            
            # Read from row 2 onwards
            for r_idx in range(2, ws.max_row + 1):
                mobile_id = ws.cell(row=r_idx, column=4).value
                if mobile_id:
                    mobile_str = str(mobile_id).strip()
                    contrato = ws.cell(row=r_idx, column=1).value
                    chasis = ws.cell(row=r_idx, column=5).value
                    linea = ws.cell(row=r_idx, column=6).value
                    
                    fleet_dict[mobile_str] = {
                        "contrato": str(contrato).strip() if contrato is not None else "",
                        "chasis": str(chasis).strip() if chasis is not None else "",
                        "linea": str(linea).strip() if linea is not None else ""
                    }
            wb.close()
            if os.path.exists(flota_temp):
                os.remove(flota_temp)
        except Exception as e:
            print(f"Error loading fleet master: {e}")
        return fleet_dict

    @staticmethod
    @st.cache_data(show_spinner=False)
    def load_novedades_master():
        """
        Loads the novelty master Excel file (Maestro_novedades.xlsx) using a powershell
        subprocess copy to avoid exclusive Windows file locking errors.
        Returns a dictionary mapping cleaned novelty observation text -> {'grupo_funcion': ..., 'criticidad': ...}
        """
        import openpyxl
        import subprocess
        
        novedades_original = os.path.join(os.getcwd(), "Flota Maestra", "Maestro_novedades.xlsx")
        novedades_temp = os.path.join(os.getcwd(), "scratch", "temp_maestro_novedades.xlsx")
        
        novedades_dict = {}
        if not os.path.exists(novedades_original):
            return novedades_dict
            
        try:
            os.makedirs(os.path.dirname(novedades_temp), exist_ok=True)
            subprocess.run(
                ["powershell", "-Command", f"Copy-Item '{novedades_original}' '{novedades_temp}'"],
                check=True,
                capture_output=True
            )
            
            wb = openpyxl.load_workbook(novedades_temp, read_only=True, data_only=True)
            ws = wb.active
            
            # Read from row 2 onwards
            for r_idx in range(2, ws.max_row + 1):
                novedad_val = ws.cell(row=r_idx, column=1).value
                if novedad_val:
                    nov_clean = str(novedad_val).strip().upper()
                    grupo = ws.cell(row=r_idx, column=2).value
                    criticidad = ws.cell(row=r_idx, column=3).value
                    
                    # Format criticidad safely (remove decimal part if represented as float)
                    criticidad_str = ""
                    if criticidad is not None:
                        try:
                            if isinstance(criticidad, float) and criticidad.is_integer():
                                criticidad_str = str(int(criticidad))
                            else:
                                criticidad_str = str(criticidad).strip()
                        except Exception:
                            criticidad_str = str(criticidad).strip()
                            
                    novedades_dict[nov_clean] = {
                        "grupo_funcion": str(grupo).strip() if grupo is not None else "",
                        "criticidad": criticidad_str
                    }
            wb.close()
            if os.path.exists(novedades_temp):
                os.remove(novedades_temp)
        except Exception as e:
            print(f"Error loading novedades master: {e}")
        return novedades_dict

    @staticmethod
    def generate_csv(df_filtered):
        """
        Generates a CSV byte stream representing the complete report,
        excluding internal columns.
        """
        df_export = df_filtered.copy()
        cols_to_drop = ['id', 'Patio', 'Técnico', 'Fecha Registro']
        for col in cols_to_drop:
            if col in df_export.columns:
                df_export = df_export.drop(columns=[col])
            
        col_mapping = {
            'Móvil': 'MÓVIL',
            'Revisión de Fugas': "REVISIÓN DE FUGAS\n(MOTOR, TRANSMISIÓN,\nDIFERENCIAL, DIRECCIÓN)",
            'Separador de Humedad': 'SEPARADOR DE HUMEDAD DEL COMBUSTIBLE',
            'Inspección Embrague': "INSPECCIÓN SISTEMA DE\nEMBRAGUE - LINEAS DEL\nCIRCUITO Y PEDAL DE MANDO",
            'Inspección Palanca': "INSPECCIÓN FUNCIÓN DE\nPALANCA DE VELOCIDADES,\nGUAYAS DE LA TRANSMISIÓN",
            'Inspección Ventilador': 'INSPECCIÓN VISUAL DEL VENTILADOR Y CORREAS',
            'Inspección Suspensión': "INSPECCIÓN SISTEMA DE\nSUSPENSIÓN DELANTERA-\nTRASERA",
            'Inspección Frenos': "FRENOS REVISIÓN\nBANDAS, CUBOS, TUBERIAS DE\nAIRE Y CONEXIONES",
            'Inspección Dirección': "INSPECCIÓN SISTEMA DE\nDIRECCIÓN\n(CAJA, BOMBA Y TERMINALES)",
            'Elec. Luces': "VERIFICAR LUCES EXTERNAS\nE INTERNAS, PITO, LIMPIABRISAS,\nDESEMPAÑADOR, ESTACIONARIAS",
            'Elec. Tablero': "VERIFICAR TABLERO DE\nINSTRUMENTOS Y MULTIPLEXADO\n(MANOMETROS, GASOMETRO, TESTIGOS)",
            'Elec. Rutero': "VERIFICAR RUTERO E\nINFORMADORES",
            'Elec. Arranque': "VERIFICAR FUNCIONAMIENTO\nDE ARRANQUE Y ALTERNADOR",
            'Elec. Puertas': "VERIFICAR APERTURA Y\nCIERRE DE PUERTAS EXTERIOR\nE INTERIOR",
            'Elec. Baterías': "VERIFICAR SUJECIÓN BATERIAS\nY ESTADO BORNES",
            'Detalles / Novedades': '          ACTIVIDADES ADICIONALES EJECUTADAS          ',
            'Observaciones': 'OBSERVACIÓN\n(CONSIGNAR CORRECCIONES Y NOVEDADES PARA PROGRAMAR)',
            'Detalles Eléctrico': '          ACTIVIDADES ADICIONALES ELÉCTRICO          ',
            'Observaciones Eléctrico': 'OBSERVACIÓN ELÉCTRICO\n(CONSIGNAR CORRECCIONES Y NOVEDADES PARA PROGRAMAR)',
            'Insumos / SAP': 'INSUMOS INSTALADOS CODIGO SAP/ CANTIDAD'
        }
        df_export = df_export.rename(columns=col_mapping)
        return df_export.to_csv(index=False).encode('utf-8')

    @staticmethod
    @st.cache_data(show_spinner=False)
    def generate_xlsx(df_filtered, patio_select, start_date, report_type, default_tecnico=""):
        """
        Generates Excel bytes utilizing official template spreadsheets based on report_type:
        - "niveles": FORMATO NIVELES.xlsx (Data starts at row 10, A-L)
        - "mecanico": FORMATO ALISTAMIENTO MECANICO.xlsx (Data starts at row 9, A-L)
        - "electrico": FORMATO ALISTAMIENTO ELECTRICO.xlsx (Data starts at row 9, A-J)
        If multiple technicians exist in the dataset, it generates a separate tab for each technician.
        """
        # Dynamic import of openpyxl to optimize app load performance
        import openpyxl
        from openpyxl.styles import Alignment, Font, Border, Side

        output = io.BytesIO()
        df_export = df_filtered.copy()
        
        # 1. Date parsing and grouping by date and technician
        df_export['Fecha_Parsed'] = pd.to_datetime(df_export['Fecha Registro'], errors='coerce')
        fallback_dt = start_date if start_date else datetime.date.today()
        df_export['Fecha_Only'] = df_export['Fecha_Parsed'].dt.date.fillna(fallback_dt)
        
        patio_val = patio_select if patio_select != "Todos los Patios" else ""

        # Map template filenames
        template_map = {
            "niveles": "FORMATO NIVELES.xlsx",
            "mecanico": "FORMATO ALISTAMIENTO MECANICO.xlsx",
            "electrico": "FORMATO ALISTAMIENTO ELECTRICO.xlsx"
        }
        template_filename = template_map.get(report_type)
        if not template_filename:
            raise ValueError(f"Unknown report type: {report_type}")

        template_path = os.path.join(
            os.getcwd(),
            "Formatos_Alistamiento",
            template_filename
        )
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template not found at: {template_path}")
        
        wb = openpyxl.load_workbook(template_path)
        ws_template = wb.active

        # Standard borders
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 2. Extract groups of (Date, Technician)
        groups = []
        if not df_export.empty:
            # Group by both Date and Technician
            grouped = df_export.groupby(['Fecha_Only', 'Técnico'])
            for (f_date, tec), df_group in grouped:
                tec_str = str(tec).strip()
                if not tec_str:
                    tec_str = default_tecnico
                groups.append((f_date, tec_str, df_group))

        if not groups:
            groups = [(fallback_dt, default_tecnico, df_export)]

        use_multiple_sheets = len(groups) > 1

        # Keep track of generated sheets so we can clean up ws_template if we created copies
        sheets_to_process = []
        
        if use_multiple_sheets:
            for f_date, tec, df_group in groups:
                ws_copy = wb.copy_worksheet(ws_template)
                # Formulate a safe, unique sheet title (max 30 chars, remove special characters)
                date_str = f"{f_date.day:02d}-{f_date.month:02d}"
                safe_title = f"{date_str} - {tec}"[:30]
                for char in ['\\', '*', '?', '/', ':', '[', ']']:
                    safe_title = safe_title.replace(char, '')
                ws_copy.title = safe_title
                
                # Copy drawings/images to copied worksheets
                for img in ws_template._images:
                    try:
                        from openpyxl.drawing.image import Image
                        if hasattr(img.ref, 'getvalue'):
                            raw_bytes = img.ref.getvalue()
                            new_stream = io.BytesIO(raw_bytes)
                            img_copy = Image(new_stream)
                        else:
                            from copy import copy
                            img_copy = copy(img)
                        img_copy.anchor = img.anchor
                        ws_copy.add_image(img_copy)
                    except Exception:
                        pass
                
                sheets_to_process.append((f_date, tec, df_group, ws_copy))
        else:
            f_date, tec, df_group = groups[0]
            if tec:
                date_str = f"{f_date.day:02d}-{f_date.month:02d}"
                safe_title = f"{date_str} - {tec}"[:30]
                for char in ['\\', '*', '?', '/', ':', '[', ']']:
                    safe_title = safe_title.replace(char, '')
                ws_template.title = safe_title
            sheets_to_process.append((f_date, tec, df_group, ws_template))

        # Process each sheet
        for f_date, tec, df_group, ws in sheets_to_process:
            # Set the exact date for this sheet's header
            day_val = f"{f_date.day:02d}"
            month_val = f"{f_date.month:02d}"
            year_val = f"{f_date.year:04d}"
            
            df_tec_export = df_group.copy()

            if report_type == "niveles":
                # Header cells
                ws['B5'] = day_val
                ws['C5'] = month_val
                ws['D5'] = year_val
                ws['F5'] = tec.upper() if tec else ""
                ws['B6'] = patio_val.upper()

                DATA_START_ROW = 10
                cols_to_keep = [
                    'Móvil', 
                    'Nivel Aceite Motor (SI/NO)', 'Nivel Aceite Motor (L)',
                    'Nivel Refrigerante (SI/NO)', 'Nivel Refrigerante (L)',
                    'Nivel Aceite Hidráulico (SI/NO)', 'Nivel Aceite Hidráulico (L)',
                    'Nivel Limpiabrisas (SI/NO)', 'Nivel Limpiabrisas (L)',
                    'Inspección Ducto Admisión', 'Drenar Tanques',
                    'Observaciones Niveles'
                ]
                cols_existing = [c for c in cols_to_keep if c in df_tec_export.columns]
                df_data = df_tec_export[cols_existing].copy()

                center_cols = {1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11}
                max_col_ref = "L"

            elif report_type == "mecanico":
                # Header cells
                ws['B5'] = day_val
                ws['C5'] = month_val
                ws['D5'] = year_val
                ws['G5'] = tec.upper() if tec else ""
                ws['B6'] = patio_val.upper()

                DATA_START_ROW = 9
                cols_to_keep = [
                    'Móvil', 'Revisión de Fugas', 'Separador de Humedad', 'Inspección Embrague',
                    'Inspección Palanca', 'Inspección Ventilador', 'Inspección Suspensión',
                    'Inspección Frenos', 'Inspección Dirección', 'Detalles / Novedades',
                    'Observaciones', 'Insumos / SAP'
                ]
                # Filter to only keep rows with actual mechanical observations
                df_tec_export = df_tec_export[
                    df_tec_export['Observaciones'].astype(str).str.strip().str.upper().isin(['', 'N/A', 'NONE', 'NAN']) == False
                ]
                cols_existing = [c for c in cols_to_keep if c in df_tec_export.columns]
                df_data = df_tec_export[cols_existing].copy()

                center_cols = {1, 2, 3, 4, 5, 6, 7, 8, 9}
                max_col_ref = "L"

            else: # electrico
                # Header cells
                ws['B5'] = day_val
                ws['C5'] = month_val
                ws['D5'] = year_val
                ws['G5'] = tec.upper() if tec else ""
                ws['B6'] = patio_val.upper()

                DATA_START_ROW = 9
                cols_to_keep = [
                    'Móvil', 'Elec. Luces', 'Elec. Tablero', 'Elec. Rutero',
                    'Elec. Arranque', 'Elec. Puertas', 'Elec. Baterías',
                    'Detalles Eléctrico', 'Observaciones Eléctrico', 'Insumos / SAP'
                ]
                # Filter to only keep rows with actual electrical observations
                df_tec_export = df_tec_export[
                    df_tec_export['Observaciones Eléctrico'].astype(str).str.strip().str.upper().isin(['', 'N/A', 'NONE', 'NAN']) == False
                ]
                cols_existing = [c for c in cols_to_keep if c in df_tec_export.columns]
                df_data = df_tec_export[cols_existing].copy()

                center_cols = {1, 2, 3, 4, 5, 6, 7}
                max_col_ref = "J"

            # Clean existing placeholders in template from DATA_START_ROW onwards
            if ws.max_row >= DATA_START_ROW:
                for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
                    for cell in row:
                        cell.value = None

            # Determine row heights & fonts
            orig_row_height = ws.row_dimensions[DATA_START_ROW].height
            if orig_row_height is None:
                orig_row_height = 42.0

            original_fonts = {}
            for col_idx in range(1, len(cols_existing) + 1):
                orig_font = ws.cell(row=DATA_START_ROW, column=col_idx).font
                if orig_font:
                    original_fonts[col_idx] = Font(
                        name=orig_font.name or 'Arial',
                        size=(orig_font.size or 10.0) + 3,
                        bold=orig_font.bold,
                        italic=orig_font.italic,
                        color=orig_font.color
                    )
                else:
                    original_fonts[col_idx] = Font(name='Arial', size=13.0)

            # Write data rows
            for row_offset, (_, data_row) in enumerate(df_data.iterrows()):
                excel_row = DATA_START_ROW + row_offset
                ws.row_dimensions[excel_row].height = orig_row_height + 3
                for col_idx, col_name in enumerate(cols_existing, start=1):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.value = data_row[col_name] if col_name in data_row and pd.notna(data_row[col_name]) else ""
                    cell.border = thin_border
                    cell.font = original_fonts[col_idx]
                    if col_idx in center_cols:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Delete extra template rows that were not filled
            next_row = DATA_START_ROW + len(df_data)
            if ws.max_row >= next_row:
                ws.delete_rows(next_row, ws.max_row - next_row + 1)

            # Set sheet properties and page setup
            ws.auto_filter.ref = f"A{DATA_START_ROW-1}:{max_col_ref}{DATA_START_ROW + len(df_data) - 1}" if len(df_data) > 0 else f"A{DATA_START_ROW-1}:{max_col_ref}{DATA_START_ROW-1}"
            ws.print_area = None
            ws.print_options.verticalCentered = False
            
            if report_type == "niveles":
                ws.print_title_rows = '1:9'
                ws.freeze_panes = 'A10'
            else:
                ws.print_title_rows = '1:8'
                ws.freeze_panes = 'A9'

        # Delete any sheets in the workbook that we did not process/write to
        processed_sheet_names = [ws.title for _, _, _, ws in sheets_to_process]
        for sheet_name in list(wb.sheetnames):
            if sheet_name not in processed_sheet_names:
                wb.remove(wb[sheet_name])

        wb.save(output)
        return output.getvalue()

    @staticmethod
    @st.cache_data(show_spinner=False)
    def generate_zip_report(df_filtered, patio_select, start_date, default_tecnico=""):
        """
        Generates a ZIP file bytes containing the three Excel reports (niveles, mecanico, electrico).
        """
        import zipfile
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            # Generate Niveles
            xlsx_niv = ExportService.generate_xlsx(df_filtered, patio_select, start_date, "niveles", default_tecnico)
            zip_file.writestr(
                f"alistamiento_niveles_CEXP_{patio_select.replace(' ', '_')}_{datetime.date.today()}.xlsx",
                xlsx_niv
            )
            
            # Generate Mecanico
            xlsx_mec = ExportService.generate_xlsx(df_filtered, patio_select, start_date, "mecanico", default_tecnico)
            zip_file.writestr(
                f"alistamiento_mecanico_CEXP_{patio_select.replace(' ', '_')}_{datetime.date.today()}.xlsx",
                xlsx_mec
            )
            
            # Generate Electrico
            xlsx_elec = ExportService.generate_xlsx(df_filtered, patio_select, start_date, "electrico", default_tecnico)
            zip_file.writestr(
                f"alistamiento_electrico_CEXP_{patio_select.replace(' ', '_')}_{datetime.date.today()}.xlsx",
                xlsx_elec
            )
            
        return zip_buffer.getvalue()

    @staticmethod
    @st.cache_data(show_spinner=False)
    def generate_novedades_volvo_xlsx(df_filtered):
        """
        Generates Excel bytes for Volvo novelty reports using Formatos_Alistamiento/FORMATO NOVEDADES VOLVO.xlsx template.
        """
        import openpyxl
        import io
        import os
        from openpyxl.styles import Alignment, Font, Border, Side
        
        output = io.BytesIO()
        template_filename = "FORMATO NOVEDADES VOLVO.xlsx"
        template_path = os.path.join(
            os.getcwd(),
            "Formatos_Alistamiento",
            template_filename
        )
        
        if not os.path.exists(template_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Novedades"
        else:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
        expected_headers = [
            "FECHA NOVEDAD", "ID", "CHASIS", "LINEA", "CONTRATO", 
            "FUENTE DE INFORMACION", "NOVEDAD", "OBSERVACIONES", 
            "GRUPO FUNCION", "CRITICIDAD", "DIAS", "ESTADO", "FECHA CORRECCION",
            "TECNICO QUE HACE CORRECCION", "INSUMOS / PARTE NUMEROS", "CANTIDAD"
        ]
        
        # Check headers (which are at row 5 in the template)
        headers = [cell.value for cell in ws[5]] if ws.max_row >= 5 else []
        headers_clean = [str(h).strip().upper() if h is not None else "" for h in headers]
        
        if not any(headers_clean):
            for col_num, h_name in enumerate(expected_headers, 1):
                ws.cell(row=5, column=col_num, value=h_name)
            headers_clean = [h.upper() for h in expected_headers]
            
        col_map = {}
        for col_idx, h_clean in enumerate(headers_clean, 1):
            if h_clean:
                col_map[h_clean] = col_idx
                
        for h_name in expected_headers:
            h_upper = h_name.upper()
            if h_upper not in col_map:
                next_col = len(col_map) + 1
                ws.cell(row=5, column=next_col, value=h_name)
                col_map[h_upper] = next_col
                
        # Clean existing data rows from row 6 onwards (preserving headers in row 5 and logo/title in rows 1-4)
        if ws.max_row >= 6:
            ws.delete_rows(6, ws.max_row - 5)
            
        # Standard borders and styling
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        font_body = Font(name='Arial', size=11)
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        fleet_info = ExportService.load_fleet_master()
        novedades_master = ExportService.load_novedades_master()
        
        for idx, row in df_filtered.iterrows():
            next_row = ws.max_row + 1
            
            f_nov = row.get("FECHA NOVEDAD", "")
            if isinstance(f_nov, (datetime.date, datetime.datetime)):
                f_nov = f_nov.strftime("%Y-%m-%d")
            elif isinstance(f_nov, str) and len(f_nov) > 10:
                f_nov = f_nov[:10]
                
            movil_id = row.get("ID", "")
            fleet_row = fleet_info.get(str(movil_id).strip(), {})
            
            chasis_val = row.get("CHASIS", "")
            if not chasis_val or str(chasis_val).strip() == "":
                chasis_val = fleet_row.get("chasis", "")
                
            linea_val = row.get("LINEA", "")
            if not linea_val or str(linea_val).strip() == "":
                linea_val = fleet_row.get("linea", "")
                
            contrato_val = row.get("CONTRATO", "")
            if not contrato_val or str(contrato_val).strip() == "":
                contrato_val = fleet_row.get("contrato", "")
                
            nov_text = row.get("NOVEDAD", "")
            nov_clean = str(nov_text).strip().upper() if nov_text else ""
            nov_master_row = novedades_master.get(nov_clean, {})
            
            grupo_val = row.get("GRUPO FUNCION", "")
            if not grupo_val or str(grupo_val).strip() == "":
                grupo_val = nov_master_row.get("grupo_funcion", "")
                
            criticidad_val = row.get("CRITICIDAD", "")
            if not criticidad_val or str(criticidad_val).strip() == "":
                criticidad_val = nov_master_row.get("criticidad", "")
                
            ws.cell(row=next_row, column=col_map["FECHA NOVEDAD"], value=f_nov)
            ws.cell(row=next_row, column=col_map["ID"], value=movil_id)
            ws.cell(row=next_row, column=col_map["CHASIS"], value=chasis_val)
            ws.cell(row=next_row, column=col_map["LINEA"], value=linea_val)
            ws.cell(row=next_row, column=col_map["CONTRATO"], value=contrato_val)
            ws.cell(row=next_row, column=col_map["FUENTE DE INFORMACION"], value=row.get("FUENTE DE INFORMACION", ""))
            ws.cell(row=next_row, column=col_map["NOVEDAD"], value=nov_text)
            ws.cell(row=next_row, column=col_map["OBSERVACIONES"], value=row.get("OBSERVACIONES", ""))
            ws.cell(row=next_row, column=col_map["GRUPO FUNCION"], value=grupo_val)
            ws.cell(row=next_row, column=col_map["CRITICIDAD"], value=criticidad_val)
            
            # Formulate DIAS dynamically based on Excel formulas
            from openpyxl.utils import get_column_letter
            col_fecha_nov_letter = get_column_letter(col_map["FECHA NOVEDAD"])
            col_estado_letter = get_column_letter(col_map["ESTADO"])
            dias_formula = f'=IF({col_estado_letter}{next_row}="CORREGIDA", 0, TODAY()-{col_fecha_nov_letter}{next_row})'
            ws.cell(row=next_row, column=col_map["DIAS"], value=dias_formula)
            ws.cell(row=next_row, column=col_map["ESTADO"], value=row.get("ESTADO", ""))
            ws.cell(row=next_row, column=col_map["FECHA CORRECCION"], value=row.get("FECHA CORRECCION", ""))
            ws.cell(row=next_row, column=col_map["TECNICO QUE HACE CORRECCION"], value=row.get("TECNICO QUE HACE CORRECCION", ""))
            ws.cell(row=next_row, column=col_map["INSUMOS / PARTE NUMEROS"], value=row.get("INSUMOS / PARTE NUMEROS", ""))
            ws.cell(row=next_row, column=col_map["CANTIDAD"], value=row.get("CANTIDAD", ""))
            
            for c_idx in col_map.values():
                cell = ws.cell(row=next_row, column=c_idx)
                cell.border = thin_border
                cell.font = font_body
                if c_idx in [col_map["FECHA NOVEDAD"], col_map["ID"], col_map["DIAS"], col_map["ESTADO"], col_map["FECHA CORRECCION"]]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
        wb.save(output)
        return output.getvalue()
